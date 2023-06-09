"""
excel文件操作
"""


from openpyxl import load_workbook

from common.param_replace import replace_label


class ExcelHandler:
    """excel 封装"""

    # 项目来说可能不变
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def open(self):
        self.wb = load_workbook(self.file_name)
        if isinstance(self.sheet_name, int):
            self.sheet = self.wb.worksheets[self.sheet_name]
        self.sheet = self.wb[self.sheet_name]


    def headers(self):
        """获取行"""
        self.open()
        # 列表推导式 拿到第一列的数据作为字典的key
        return [cell.value for cell in self.sheet[1]]

    def read(self, start_row=2, start_column=1, dict_flag=True):
        """
        获取所有的数据
        :param start_row:   开始行
        :param start_column:  开始列
        :param dict_flag:  返回字典格式标志
        :return:  返回数据
        """
        self.open()
        # 获取herders
        title = self.headers()
        data = []
        for row in range(start_row, self.sheet.max_row + 1):
            row_data = []
            for column in range(start_column, self.sheet.max_column + 1):
                row_data.append(self.sheet.cell(row, column).value)
            if dict_flag:  # 将title 和 每行的数据列表转换为字典格式返回
                data.append(dict(zip(title, row_data)))
            else:  # 返回列表格式数据
                data.append(row_data)
        return data

    def read_cell(self, row, column):
        """一个单元格的数据"""
        self.open()
        return self.sheet.cell(row, column).value

    def save(self):
        """保存"""
        self.wb.save(self.file_name)
        self.wb.close()

    def write(self, row, column, data):
        self.open()
        self.sheet.cell(row, column).value = data
        # 保存关闭
        self.wb.save(self.file_name)
        self.wb.close()


if __name__ == '__main__':
    xls = ExcelHandler(r'../data/test_case1.xlsx', 'check_work')
    # print(xls.read(dict_flag=False))
    data = xls.read()
    print(data)
    for ele in data:
        data = ele["data"]
        print(replace_label(data))

